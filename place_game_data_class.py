# Author : Charlie Newell
# Date   : 11/14/2017
""" The purpose of this file is to take the game data from the driver class and place it properly in an 
    excel document
"""

# Excel manipulation libraries
import openpyxl
from openpyxl import load_workbook


# function to open workbook
def openBook(filename, sheetname):

    # open proper workbook and worksheet, get max row of sheet
    wb = load_workbook(filename)
    ws = wb.get_sheet_by_name(sheetname)
    maxRow = ws.max_row

    return wb, ws, maxRow

# function to place on games data in excel document
def placeData(wb, ws, winnerData, loserData, maxRow):

    # place winning team data
    # ws.cell(row=1, column=5).value = winnerData['Team']
    ws['F' + str(maxRow)] = winnerData['Team']
    ws['G' + str(maxRow)] = winnerData['PTS']
    ws['J' + str(maxRow)] = winnerData['FG']
    ws['K' + str(maxRow)] = winnerData['FGA']
    ws['L' + str(maxRow)] = winnerData['FGPer']
    ws['M' + str(maxRow)] = winnerData['TP']
    ws['N' + str(maxRow)] = winnerData['TPA']
    ws['O' + str(maxRow)] = winnerData['TPPer']
    ws['P' + str(maxRow)] = winnerData['FT']
    ws['Q' + str(maxRow)] = winnerData['FTA']
    ws['R' + str(maxRow)] = winnerData['FTPer']
    ws['S' + str(maxRow)] = winnerData['ORB']
    ws['T' + str(maxRow)] = winnerData['TRB']
    ws['U' + str(maxRow)] = winnerData['AST']
    ws['V' + str(maxRow)] = winnerData['STL']
    ws['W' + str(maxRow)] = winnerData['BLK']
    ws['X' + str(maxRow)] = winnerData['TOV']
    ws['Y' + str(maxRow)] = winnerData['PF']

    wb.save('test.xlsx')

    

